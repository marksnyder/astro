"""FastAPI backend for Astro web UI."""

import shutil
import tempfile
from typing import Optional

from datetime import datetime, timezone

import fastapi
from fastapi import FastAPI, HTTPException, UploadFile, WebSocket, WebSocketDisconnect
from fastapi.middleware.cors import CORSMiddleware
from fastapi.responses import FileResponse, HTMLResponse
from fastapi.staticfiles import StaticFiles
from pathlib import Path
from pydantic import BaseModel

from dotenv import load_dotenv

load_dotenv()

from src.ingest import SUPPORTED_EXTENSIONS, chunk_documents, load_document
from src.markdowns import (
    FEED_FILES_DIR,
    IMAGES_DIR,
    agent_task_to_dict,
    add_markdown_image,
    category_to_dict,
    create_agent_task,
    create_category,
    create_feed,
    create_feed_post_file,
    create_feed_post_markdown,
    list_pinned_feeds,
    list_pinned_categories,
    create_link,
    create_markdown,
    create_universe,
    delete_agent_task,
    delete_all_markdown_images,
    delete_category,
    delete_document_meta,
    delete_feed,
    delete_feed_post,
    delete_link,
    delete_markdown,
    delete_markdown_image,
    delete_universe,
    feed_post_to_dict,
    feed_to_dict,
    get_agent_task,
    get_all_document_categories,
    get_all_document_meta,
    get_document_paths_for_category,
    get_document_pinned,
    get_feed,
    get_feed_post,
    set_feed_pinned,
    set_category_pinned,
    get_link,
    get_markdown,
    get_universe,
    get_universe_document_paths,
    get_universe_markdown_ids,
    list_agent_tasks,
    list_feed_posts,
    list_feed_posts_by_category,
    mark_feed_posts_read,
    get_unread_counts_by_category,
    get_recent_counts_by_category,
    list_feeds,
    list_links,
    list_categories,
    list_markdown_images,
    list_markdowns,
    list_pinned_documents,
    list_pinned_links,
    list_pinned_markdowns,
    category_in_universe,
    list_universes,
    link_to_dict,
    markdown_image_to_dict,
    markdown_to_dict,
    move_category,
    move_diagram_to_universe,
    move_feed_to_universe,
    move_link_to_universe,
    move_markdown_to_universe,
    move_table_to_universe,
    update_category,
    update_feed,
    rename_universe,
    set_document_category,
    set_document_pinned,
    set_document_universe,
    set_link_pinned,
    set_markdown_pinned,
    get_setting,
    set_setting,
    universe_to_dict,
    update_agent_task,
    update_link,
    update_markdown,
    list_post_comments,
    create_post_comment,
    update_post_comment,
    delete_post_comment,
    post_comment_to_dict,
    create_diagram,
    create_table,
    create_table_row,
    delete_diagram,
    delete_table,
    delete_table_row,
    diagram_summary_to_dict,
    diagram_to_dict,
    get_diagram,
    get_table,
    get_table_row,
    list_diagram_summaries,
    list_pinned_diagram_summaries,
    list_pinned_tables,
    list_table_rows,
    list_tables,
    set_diagram_pinned,
    set_table_pinned,
    table_row_to_dict,
    table_to_dict,
    update_diagram,
    update_table,
    update_table_row,
)
from src.store import (
    add_documents,
    delete_document_chunks,
    delete_markdown_from_store,
    doc_count,
    get_retriever,
    upsert_markdown,
)

DOCUMENTS_DIR = Path(__file__).resolve().parent.parent / "documents"

_ngircd_proc = None


def _start_ngircd():
    """Launch ngircd as a background process if not already running."""
    global _ngircd_proc
    import subprocess, shutil, time
    if shutil.which("ngircd") is None:
        print("[IRC] ngircd not found — install it with: apt install ngircd")
        return
    conf = Path(__file__).resolve().parent.parent / "config" / "ngircd.conf"
    if not conf.exists():
        print(f"[IRC] Config not found: {conf}")
        return
    try:
        _ngircd_proc = subprocess.Popen(
            ["ngircd", "-n", "-f", str(conf)],
            stdout=subprocess.DEVNULL,
            stderr=subprocess.DEVNULL,
        )
        print(f"[IRC] ngircd started (config: {conf})")
        time.sleep(0.5)
    except Exception as e:
        print(f"[IRC] Failed to start ngircd: {e}")


# MCP server for AI agent integration
from contextlib import asynccontextmanager
from src.mcp_server import mcp as _mcp

_mcp_app = _mcp.http_app(path="/", stateless_http=True)


@asynccontextmanager
async def _lifespan(application):
    _start_ngircd()
    from src.irc_client import IRCClient
    IRCClient.get()
    from src.irc_monitor import IRCMonitor
    IRCMonitor.get()
    from src.agent_task_runner import AgentTaskRunner
    AgentTaskRunner.get()

    async with _mcp_app.lifespan(application):
        yield

    from src.irc_client import IRCClient as _IC
    if _IC._instance:
        _IC._instance.stop()
    from src.irc_monitor import IRCMonitor as _IM
    if _IM._instance:
        _IM._instance.stop()
    from src.agent_task_runner import AgentTaskRunner as _ATR
    if _ATR._instance is not None:
        _ATR._instance.stop()
    if _ngircd_proc:
        _ngircd_proc.terminate()
        try:
            _ngircd_proc.wait(timeout=3)
        except Exception:
            _ngircd_proc.kill()


app = FastAPI(title="Astro", version="1.0.0", lifespan=_lifespan)

app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],
    allow_methods=["*"],
    allow_headers=["*"],
)

app.mount("/mcp", _mcp_app)


@app.middleware("http")
async def api_key_middleware(request: fastapi.Request, call_next):
    path = request.url.path
    skip_paths = ("/api/auth/", "/api/version", "/assets/", "/index.html", "/favicon")
    if path == "/" or any(path.startswith(p) for p in skip_paths):
        return await call_next(request)
    if not path.startswith("/api/") and not path.startswith("/mcp"):
        return await call_next(request)
    stored_key = get_setting("api_key", "")
    if not stored_key:
        return await call_next(request)
    provided = request.headers.get("x-api-key", "") or request.query_params.get("api_key", "")
    if provided == stored_key:
        return await call_next(request)
    from fastapi.responses import JSONResponse
    return JSONResponse(status_code=401, content={"detail": "Invalid or missing API key"})



# ── Schemas ───────────────────────────────────────────────────────────────



class StatsResponse(BaseModel):
    chunks: int
    schema_version: int = 0


class MarkdownRequest(BaseModel):
    title: str
    body: str
    category_id: Optional[int] = None


class MarkdownResponse(BaseModel):
    id: int
    title: str
    body: str
    category_id: Optional[int]
    pinned: bool
    created_at: str
    updated_at: str
    universe_id: int = 1


class MoveToUniverseRequest(BaseModel):
    universe_id: int
    category_id: Optional[int] = None


class CategoryRequest(BaseModel):
    name: str
    parent_id: Optional[int] = None
    emoji: Optional[str] = None


class CategoryUpdateRequest(BaseModel):
    name: str
    emoji: Optional[str] = None


class CategoryResponse(BaseModel):
    id: int
    name: str
    parent_id: Optional[int]
    universe_id: int = 1
    emoji: Optional[str] = None
    sort_order: int = 0


class DocumentInfo(BaseModel):
    name: str
    path: str
    folder: str
    size: int
    extension: str
    category_id: Optional[int]
    pinned: bool = False
    modified_at: str = ""


class DocumentCategoryRequest(BaseModel):
    category_id: Optional[int] = None


class LinkRequest(BaseModel):
    title: str
    url: str
    category_id: Optional[int] = None


class LinkResponse(BaseModel):
    id: int
    title: str
    url: str
    category_id: Optional[int]
    pinned: bool
    created_at: str
    updated_at: str
    universe_id: int = 1


# ── Universes ─────────────────────────────────────────────────────────────


class UniverseRequest(BaseModel):
    name: str


class UniverseResponse(BaseModel):
    id: int
    name: str
    created_at: str
    updated_at: str


@app.get("/api/universes", response_model=list[UniverseResponse])
def api_list_universes():
    return [universe_to_dict(u) for u in list_universes()]


@app.post("/api/universes", response_model=UniverseResponse, status_code=201)
def api_create_universe(req: UniverseRequest):
    u = create_universe(req.name.strip())
    return universe_to_dict(u)


@app.put("/api/universes/{uid}", response_model=UniverseResponse)
def api_rename_universe(uid: int, req: UniverseRequest):
    u = rename_universe(uid, req.name.strip())
    if not u:
        raise HTTPException(status_code=404, detail="Universe not found")
    return universe_to_dict(u)


@app.delete("/api/universes/{uid}")
def api_delete_universe(uid: int):
    """Delete a universe and ALL its content. Cannot delete the last universe."""
    u = get_universe(uid)
    if not u:
        raise HTTPException(status_code=404, detail="Universe not found")

    # Clean up vector store entries before deleting DB rows
    for nid in get_universe_markdown_ids(uid):
        delete_markdown_from_store(nid)
    for path in get_universe_document_paths(uid):
        delete_document_chunks(str(DOCUMENTS_DIR / path))
        # Remove physical file
        f = DOCUMENTS_DIR / path
        if f.is_file():
            f.unlink()

    if not delete_universe(uid):
        raise HTTPException(status_code=400, detail="Cannot delete the last universe")
    return {"ok": True}


# ── Categories ────────────────────────────────────────────────────────────


@app.get("/api/categories", response_model=list[CategoryResponse])
def api_list_categories(universe_id: Optional[int] = None):
    return [category_to_dict(c) for c in list_categories(universe_id=universe_id)]


@app.post("/api/categories", response_model=CategoryResponse, status_code=201)
def api_create_category(req: CategoryRequest, universe_id: int = 1):
    cat = create_category(req.name, req.parent_id, universe_id=universe_id, emoji=req.emoji)
    return category_to_dict(cat)


@app.put("/api/categories/{cat_id}", response_model=CategoryResponse)
def api_update_category(cat_id: int, req: CategoryUpdateRequest):
    cat = update_category(cat_id, name=req.name, emoji=req.emoji)
    if not cat:
        raise HTTPException(status_code=404, detail="Category not found")
    return category_to_dict(cat)


@app.put("/api/categories/{cat_id}/move", response_model=CategoryResponse)
def api_move_category(cat_id: int, direction: str):
    if direction not in ("up", "down"):
        raise HTTPException(status_code=400, detail="direction must be 'up' or 'down'")
    cat = move_category(cat_id, direction)
    if not cat:
        raise HTTPException(status_code=400, detail="Cannot move category in that direction")
    return category_to_dict(cat)


@app.delete("/api/categories/{cat_id}")
def api_delete_category(cat_id: int):
    if not delete_category(cat_id):
        raise HTTPException(status_code=404, detail="Category not found")
    return {"ok": True}


@app.put("/api/categories/{cat_id}/pin")
def api_pin_category(cat_id: int, pinned: bool = True):
    if not set_category_pinned(cat_id, pinned):
        raise HTTPException(status_code=404, detail="Category not found")
    return {"ok": True}


# ── Markdowns ─────────────────────────────────────────────────────────────


@app.get("/api/markdowns", response_model=list[MarkdownResponse])
def api_list_markdowns(q: str = "", category_id: Optional[int] = None, universe_id: Optional[int] = None):
    return [markdown_to_dict(n) for n in list_markdowns(q, category_id, universe_id=universe_id)]


@app.get("/api/markdowns/{markdown_id}", response_model=MarkdownResponse)
def api_get_markdown(markdown_id: int):
    markdown = get_markdown(markdown_id)
    if not markdown:
        raise HTTPException(status_code=404, detail="Markdown not found")
    return markdown_to_dict(markdown)


@app.post("/api/markdowns", response_model=MarkdownResponse, status_code=201)
def api_create_markdown(req: MarkdownRequest, universe_id: int = 1):
    markdown = create_markdown(req.title, req.body, req.category_id, universe_id=universe_id)
    try:
        upsert_markdown(markdown.id, f"{markdown.title}\n\n{markdown.body}", markdown.title, universe_id=universe_id)
    except Exception as e:
        print(f"[Astro] WARNING: Failed to upsert markdown {markdown.id} into vector store: {e}")
    return markdown_to_dict(markdown)


@app.put("/api/markdowns/{markdown_id}", response_model=MarkdownResponse)
def api_update_markdown(markdown_id: int, req: MarkdownRequest):
    markdown = update_markdown(markdown_id, req.title, req.body, req.category_id)
    if not markdown:
        raise HTTPException(status_code=404, detail="Markdown not found")
    try:
        upsert_markdown(markdown.id, f"{markdown.title}\n\n{markdown.body}", markdown.title, universe_id=markdown.universe_id)
    except Exception as e:
        print(f"[Astro] WARNING: Failed to upsert markdown {markdown.id} into vector store: {e}")
    return markdown_to_dict(markdown)


def _validate_move_category(req: MoveToUniverseRequest) -> None:
    if not category_in_universe(req.category_id, req.universe_id):
        raise HTTPException(status_code=400, detail="Category does not belong to the target universe")


@app.post("/api/markdowns/{markdown_id}/move-universe", response_model=MarkdownResponse)
def api_move_markdown_universe(markdown_id: int, req: MoveToUniverseRequest):
    if not get_universe(req.universe_id):
        raise HTTPException(status_code=400, detail="Universe not found")
    _validate_move_category(req)
    markdown = move_markdown_to_universe(markdown_id, req.universe_id, req.category_id)
    if not markdown:
        raise HTTPException(status_code=404, detail="Markdown not found")
    return markdown_to_dict(markdown)


@app.delete("/api/markdowns/{markdown_id}")
def api_delete_markdown(markdown_id: int):
    delete_all_markdown_images(markdown_id)
    if not delete_markdown(markdown_id):
        raise HTTPException(status_code=404, detail="Markdown not found")
    delete_markdown_from_store(markdown_id)
    return {"ok": True}


# ── Pinned items ──────────────────────────────────────────────────────────


@app.put("/api/markdowns/{markdown_id}/pin")
def api_toggle_markdown_pin(markdown_id: int, pinned: bool = True):
    if not set_markdown_pinned(markdown_id, pinned):
        raise HTTPException(status_code=404, detail="Markdown not found")
    return {"ok": True}


@app.put("/api/documents/pin")
def api_toggle_document_pin(path: str, pinned: bool = True):
    safe = (DOCUMENTS_DIR / path).resolve()
    if not str(safe).startswith(str(DOCUMENTS_DIR)) or not safe.is_file():
        raise HTTPException(status_code=404, detail="File not found")
    set_document_pinned(path, pinned)
    return {"ok": True}


@app.get("/api/pinned")
def api_list_pinned(universe_id: Optional[int] = None):
    """Return all pinned markdowns, documents, and links in one call."""
    markdowns = [markdown_to_dict(m) for m in list_pinned_markdowns(universe_id=universe_id)]
    doc_paths = list_pinned_documents(universe_id=universe_id)
    docs = []
    for rel_str in doc_paths:
        f = DOCUMENTS_DIR / rel_str
        if f.is_file():
            docs.append({
                "name": f.name,
                "path": rel_str,
                "extension": f.suffix.lower().lstrip("."),
                "size": f.stat().st_size,
            })
    links = [link_to_dict(l) for l in list_pinned_links(universe_id=universe_id)]
    feeds = [feed_to_dict(f) for f in list_pinned_feeds(universe_id=universe_id)]
    pinned_cats = [category_to_dict(c) for c in list_pinned_categories(universe_id=universe_id)]
    diagrams = [diagram_summary_to_dict(d) for d in list_pinned_diagram_summaries(universe_id=universe_id)]
    tables = [table_to_dict(t) for t in list_pinned_tables(universe_id=universe_id)]
    return {"markdowns": markdowns, "documents": docs, "links": links, "feeds": feeds, "feed_categories": pinned_cats, "diagrams": diagrams, "tables": tables}


# ── Links (bookmarks) ───────────────────────────────────────────────────


@app.get("/api/links", response_model=list[LinkResponse])
def api_list_links(q: str = "", category_id: Optional[int] = None, universe_id: Optional[int] = None):
    return [link_to_dict(l) for l in list_links(q, category_id, universe_id=universe_id)]


@app.get("/api/links/{link_id}", response_model=LinkResponse)
def api_get_link(link_id: int):
    link = get_link(link_id)
    if not link:
        raise HTTPException(status_code=404, detail="Link not found")
    return link_to_dict(link)


@app.post("/api/links", response_model=LinkResponse, status_code=201)
def api_create_link(req: LinkRequest, universe_id: int = 1):
    link = create_link(req.title, req.url, req.category_id, universe_id=universe_id)
    return link_to_dict(link)


@app.put("/api/links/{link_id}", response_model=LinkResponse)
def api_update_link(link_id: int, req: LinkRequest):
    link = update_link(link_id, req.title, req.url, req.category_id)
    if not link:
        raise HTTPException(status_code=404, detail="Link not found")
    return link_to_dict(link)


@app.post("/api/links/{link_id}/move-universe", response_model=LinkResponse)
def api_move_link_universe(link_id: int, req: MoveToUniverseRequest):
    if not get_universe(req.universe_id):
        raise HTTPException(status_code=400, detail="Universe not found")
    _validate_move_category(req)
    link = move_link_to_universe(link_id, req.universe_id, req.category_id)
    if not link:
        raise HTTPException(status_code=404, detail="Link not found")
    return link_to_dict(link)


@app.delete("/api/links/{link_id}")
def api_delete_link(link_id: int):
    if not delete_link(link_id):
        raise HTTPException(status_code=404, detail="Link not found")
    return {"ok": True}


@app.put("/api/links/{link_id}/pin")
def api_toggle_link_pin(link_id: int, pinned: bool = True):
    if not set_link_pinned(link_id, pinned):
        raise HTTPException(status_code=404, detail="Link not found")
    return {"ok": True}


# ── Markdown images ──────────────────────────────────────────────────────

IMAGE_EXTENSIONS = {".png", ".jpg", ".jpeg", ".gif", ".webp", ".bmp", ".svg"}


@app.get("/api/markdowns/{markdown_id}/images")
def api_list_markdown_images(markdown_id: int):
    return [markdown_image_to_dict(img) for img in list_markdown_images(markdown_id)]


@app.post("/api/markdowns/{markdown_id}/images", status_code=201)
async def api_upload_markdown_image(markdown_id: int, file: UploadFile):
    markdown = get_markdown(markdown_id)
    if not markdown:
        raise HTTPException(status_code=404, detail="Markdown not found")
    if not file.filename:
        raise HTTPException(status_code=400, detail="No filename provided")
    ext = Path(file.filename).suffix.lower()
    if ext not in IMAGE_EXTENSIONS:
        raise HTTPException(
            status_code=400,
            detail=f"Unsupported image type: {ext}. Supported: {', '.join(sorted(IMAGE_EXTENSIONS))}",
        )
    data = await file.read()
    img = add_markdown_image(markdown_id, file.filename, data)
    return markdown_image_to_dict(img)


@app.delete("/api/markdown-images/{image_id}")
def api_delete_markdown_image(image_id: int):
    if not delete_markdown_image(image_id):
        raise HTTPException(status_code=404, detail="Image not found")
    return {"ok": True}


@app.get("/api/markdown-images/file/{filename}")
def api_serve_markdown_image(filename: str):
    safe = (IMAGES_DIR / filename).resolve()
    if not str(safe).startswith(str(IMAGES_DIR.resolve())) or not safe.is_file():
        raise HTTPException(status_code=404, detail="Image not found")
    return FileResponse(safe)


# ── Documents (archive) ──────────────────────────────────────────────────


@app.get("/api/documents", response_model=list[DocumentInfo])
def api_list_documents(q: str = "", category_id: Optional[int] = None, universe_id: Optional[int] = None):
    """List documents, optionally filtered by search, category, and universe."""
    if not DOCUMENTS_DIR.exists():
        return []

    meta_map = get_all_document_meta(universe_id=universe_id)

    # If filtering by category, pre-compute the allowed paths
    allowed_paths: set[str] | None = None
    if category_id is not None:
        allowed_paths = get_document_paths_for_category(category_id)

    results: list[DocumentInfo] = []
    for f in DOCUMENTS_DIR.rglob("*"):
        if not f.is_file():
            continue
        if f.suffix.lower() not in SUPPORTED_EXTENSIONS:
            continue
        rel = f.relative_to(DOCUMENTS_DIR)
        rel_str = str(rel)
        folder = str(rel.parent) if str(rel.parent) != "." else ""
        if q and q.lower() not in f.name.lower():
            continue
        if allowed_paths is not None and rel_str not in allowed_paths:
            continue
        if universe_id is not None and rel_str not in meta_map:
            continue
        meta = meta_map.get(rel_str, {})
        stat = f.stat()
        mtime = datetime.fromtimestamp(stat.st_mtime, tz=timezone.utc).isoformat()
        results.append(
            DocumentInfo(
                name=f.name,
                path=rel_str,
                folder=folder,
                size=stat.st_size,
                extension=f.suffix.lower().lstrip("."),
                category_id=meta.get("category_id"),
                pinned=meta.get("pinned", False),
                modified_at=mtime,
            )
        )
    results.sort(key=lambda d: d.modified_at, reverse=True)
    return results


@app.get("/api/documents/download")
def api_download_document(path: str):
    safe = (DOCUMENTS_DIR / path).resolve()
    if not str(safe).startswith(str(DOCUMENTS_DIR)):
        raise HTTPException(status_code=400, detail="Invalid path")
    if not safe.is_file():
        raise HTTPException(status_code=404, detail="File not found")
    return FileResponse(safe, filename=safe.name)


@app.get("/api/documents/view")
def api_view_document(path: str):
    """Serve a document inline (for in-browser viewing)."""
    safe = (DOCUMENTS_DIR / path).resolve()
    if not str(safe).startswith(str(DOCUMENTS_DIR)):
        raise HTTPException(status_code=400, detail="Invalid path")
    if not safe.is_file():
        raise HTTPException(status_code=404, detail="File not found")

    ext = safe.suffix.lower()

    # XLSX / XLS → render as HTML table
    if ext in (".xlsx", ".xls"):
        from openpyxl import load_workbook
        from html import escape

        wb = load_workbook(str(safe), data_only=True)
        sheets_html = []
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            rows = list(ws.iter_rows(values_only=True))
            if not rows:
                continue
            table = f"<h2>{escape(sheet_name)}</h2><table><thead><tr>"
            # Use first row as header
            for cell in rows[0]:
                table += f"<th>{escape(str(cell)) if cell is not None else ''}</th>"
            table += "</tr></thead><tbody>"
            for row in rows[1:]:
                table += "<tr>"
                for cell in row:
                    table += f"<td>{escape(str(cell)) if cell is not None else ''}</td>"
                table += "</tr>"
            table += "</tbody></table>"
            sheets_html.append(table)

        html = f"""<!DOCTYPE html>
<html><head>
<meta charset="utf-8">
<title>{escape(safe.name)}</title>
<style>
  body {{ font-family: -apple-system, BlinkMacSystemFont, 'Segoe UI', Roboto, sans-serif;
         margin: 2rem; background: #f8f9fa; color: #1a1a2e; }}
  h2 {{ margin-top: 2rem; color: #333; }}
  table {{ border-collapse: collapse; width: 100%; margin-bottom: 2rem;
           background: #fff; box-shadow: 0 1px 3px rgba(0,0,0,.1); border-radius: 6px;
           overflow: hidden; }}
  th, td {{ border: 1px solid #e0e0e0; padding: 8px 12px; text-align: left; }}
  th {{ background: #4a5568; color: #fff; font-weight: 600; }}
  tr:nth-child(even) {{ background: #f7fafc; }}
  tr:hover {{ background: #edf2f7; }}
</style>
</head><body>
<h1>{escape(safe.name)}</h1>
{''.join(sheets_html)}
</body></html>"""
        return HTMLResponse(content=html)

    # PDF → inline
    return FileResponse(
        safe,
        media_type="application/pdf",
        headers={"Content-Disposition": f"inline; filename=\"{safe.name}\""},
    )


@app.put("/api/documents/category")
def api_set_document_category(path: str, req: DocumentCategoryRequest):
    safe = (DOCUMENTS_DIR / path).resolve()
    if not str(safe).startswith(str(DOCUMENTS_DIR)) or not safe.is_file():
        raise HTTPException(status_code=404, detail="File not found")
    set_document_category(path, req.category_id)
    return {"ok": True}


@app.post("/api/documents/move-universe")
def api_move_document_universe(path: str, req: MoveToUniverseRequest):
    """Move an archived document to another universe; optional category; re-ingests vector chunks."""
    if not get_universe(req.universe_id):
        raise HTTPException(status_code=400, detail="Universe not found")
    _validate_move_category(req)
    safe = (DOCUMENTS_DIR / path).resolve()
    if not str(safe).startswith(str(DOCUMENTS_DIR)) or not safe.is_file():
        raise HTTPException(status_code=404, detail="File not found")
    documents = load_document(str(safe))
    if not documents:
        raise HTTPException(status_code=400, detail="Could not extract content from file for re-indexing")
    delete_document_chunks(str(safe))
    for doc in documents:
        doc.metadata["source"] = str(safe)
    chunks = chunk_documents(documents)
    add_documents(chunks, universe_id=req.universe_id)
    set_document_universe(path, req.universe_id)
    set_document_category(path, req.category_id, req.universe_id)
    return {"ok": True, "path": path}


@app.delete("/api/documents")
def api_delete_document(path: str):
    safe = (DOCUMENTS_DIR / path).resolve()
    if not str(safe).startswith(str(DOCUMENTS_DIR)):
        raise HTTPException(status_code=400, detail="Invalid path")
    if not safe.is_file():
        raise HTTPException(status_code=404, detail="File not found")
    removed = delete_document_chunks(str(safe))
    rel = str(safe.relative_to(DOCUMENTS_DIR))
    delete_document_meta(rel)
    safe.unlink()
    return {"ok": True, "chunks_removed": removed}


@app.post("/api/documents/upload")
def api_upload_document(file: UploadFile, universe_id: int = 1):
    if not file.filename:
        raise HTTPException(status_code=400, detail="No filename provided")
    ext = Path(file.filename).suffix.lower()
    if ext not in SUPPORTED_EXTENSIONS:
        raise HTTPException(
            status_code=400,
            detail=f"Unsupported file type: {ext}. Supported: {', '.join(sorted(SUPPORTED_EXTENSIONS))}",
        )
    DOCUMENTS_DIR.mkdir(parents=True, exist_ok=True)
    with tempfile.NamedTemporaryFile(delete=False, suffix=ext) as tmp:
        shutil.copyfileobj(file.file, tmp)
        tmp_path = tmp.name
    try:
        documents = load_document(tmp_path)
        if not documents:
            raise HTTPException(status_code=400, detail="Could not extract content from file")
        archive_folder = ext.lstrip(".")
        archive_dir = DOCUMENTS_DIR / archive_folder
        archive_dir.mkdir(exist_ok=True)
        dest = archive_dir / file.filename
        counter = 1
        while dest.exists():
            stem = Path(file.filename).stem
            dest = archive_dir / f"{stem}_{counter}{ext}"
            counter += 1
        for doc in documents:
            doc.metadata["source"] = str(dest)
        chunks = chunk_documents(documents)
        add_documents(chunks, universe_id=universe_id)
        shutil.move(tmp_path, str(dest))
    except HTTPException:
        Path(tmp_path).unlink(missing_ok=True)
        raise
    except Exception as e:
        Path(tmp_path).unlink(missing_ok=True)
        raise HTTPException(status_code=500, detail=f"Ingestion failed: {e}")
    rel = dest.relative_to(DOCUMENTS_DIR)
    set_document_universe(str(rel), universe_id)
    return {"name": dest.name, "path": str(rel), "chunks": len(chunks)}


# ── Agent Tasks ────────────────────────────────────────────────────────────


class AgentTaskRequest(BaseModel):
    title: str
    markdown_id: int
    channel: str
    universe_id: int = 1
    schedule_mode: str = "manual"  # manual | cron | once
    cron_expr: str = ""
    run_at: str | None = None
    enabled: bool = True


def _validate_agent_task_markdown(markdown_id: int, universe_id: int) -> None:
    m = get_markdown(markdown_id)
    if not m:
        raise HTTPException(status_code=404, detail="Markdown not found")
    if m.universe_id != universe_id:
        raise HTTPException(status_code=400, detail="Markdown must belong to the selected universe")


def _normalize_channel(ch: str) -> str:
    ch = ch.strip()
    if not ch.startswith("#"):
        ch = "#" + ch
    return ch


@app.get("/api/agent-tasks")
def api_list_agent_tasks(universe_id: Optional[int] = None):
    tasks = list_agent_tasks(universe_id=universe_id)
    out = []
    for t in tasks:
        md = get_markdown(t.markdown_id)
        out.append(agent_task_to_dict(t, md.title if md else None))
    return out


@app.post("/api/agent-tasks", status_code=201)
def api_create_agent_task(req: AgentTaskRequest):
    if req.schedule_mode not in ("manual", "cron", "once"):
        raise HTTPException(status_code=400, detail="Invalid schedule_mode")
    if req.schedule_mode == "cron" and not (req.cron_expr or "").strip():
        raise HTTPException(status_code=400, detail="cron_expr required for cron schedule")
    if req.schedule_mode == "once" and not (req.run_at or "").strip():
        raise HTTPException(status_code=400, detail="run_at required for one-time schedule")
    _validate_agent_task_markdown(req.markdown_id, req.universe_id)
    ch = _normalize_channel(req.channel)
    t = create_agent_task(
        title=req.title,
        markdown_id=req.markdown_id,
        channel=ch,
        universe_id=req.universe_id,
        schedule_mode=req.schedule_mode,
        cron_expr=(req.cron_expr or "").strip() or None,
        run_at=(req.run_at or "").strip() or None,
        enabled=req.enabled,
    )
    md = get_markdown(t.markdown_id)
    return agent_task_to_dict(t, md.title if md else None)


@app.put("/api/agent-tasks/{task_id}")
def api_update_agent_task(task_id: int, req: AgentTaskRequest):
    if req.schedule_mode not in ("manual", "cron", "once"):
        raise HTTPException(status_code=400, detail="Invalid schedule_mode")
    if req.schedule_mode == "cron" and not (req.cron_expr or "").strip():
        raise HTTPException(status_code=400, detail="cron_expr required for cron schedule")
    if req.schedule_mode == "once" and not (req.run_at or "").strip():
        raise HTTPException(status_code=400, detail="run_at required for one-time schedule")
    _validate_agent_task_markdown(req.markdown_id, req.universe_id)
    ch = _normalize_channel(req.channel)
    t = update_agent_task(
        task_id,
        title=req.title,
        markdown_id=req.markdown_id,
        channel=ch,
        universe_id=req.universe_id,
        schedule_mode=req.schedule_mode,
        cron_expr=(req.cron_expr or "").strip() or None,
        run_at=(req.run_at or "").strip() or None,
        enabled=req.enabled,
    )
    if not t:
        raise HTTPException(status_code=404, detail="Task not found")
    md = get_markdown(t.markdown_id)
    return agent_task_to_dict(t, md.title if md else None)


@app.delete("/api/agent-tasks/{task_id}")
def api_delete_agent_task(task_id: int):
    if not delete_agent_task(task_id):
        raise HTTPException(status_code=404, detail="Task not found")
    return {"ok": True}


@app.post("/api/agent-tasks/{task_id}/run")
def api_run_agent_task_now(task_id: int):
    from src.agent_task_runner import ChannelCooldownError, send_agent_task_message_now

    try:
        send_agent_task_message_now(task_id)
        return {"ok": True}
    except ValueError as e:
        raise HTTPException(status_code=400, detail=str(e))
    except ChannelCooldownError as e:
        raise HTTPException(
            status_code=429,
            detail=f"Channel {e.channel} was sent to recently; wait {e.wait_seconds:.0f}s",
        )
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Failed to send: {e}")


# ── App settings ──────────────────────────────────────────────────────────


class SettingRequest(BaseModel):
    value: str


@app.get("/api/settings/{key}")
def api_get_setting(key: str):
    out: dict = {"key": key, "value": get_setting(key) or ""}
    if key == "agent_task_message_template":
        from src.agent_task_runner import DEFAULT_AGENT_TASK_TEMPLATE

        out["default_value"] = DEFAULT_AGENT_TASK_TEMPLATE
    return out


@app.put("/api/settings/{key}")
def api_set_setting(key: str, req: SettingRequest):
    set_setting(key, req.value)
    return {"ok": True}


# ── Auth ──────────────────────────────────────────────────────────────────


@app.get("/api/auth/status")
def api_auth_status():
    """Check if API key auth is enabled and if a provided key is valid."""
    stored = get_setting("api_key", "")
    return {"enabled": bool(stored)}


@app.post("/api/auth/login")
def api_auth_login(req: SettingRequest):
    """Validate an API key."""
    stored = get_setting("api_key", "")
    if not stored:
        return {"ok": True}
    if req.value == stored:
        return {"ok": True, "api_key": stored}
    raise HTTPException(status_code=401, detail="Invalid API key")


@app.post("/api/auth/generate-key")
def api_auth_generate_key():
    """Generate a new API key (requires existing key if one is set)."""
    import uuid
    new_key = str(uuid.uuid4())
    set_setting("api_key", new_key)
    return {"api_key": new_key}


@app.post("/api/auth/clear-key")
def api_auth_clear_key():
    """Clear the API key, making the app open again."""
    set_setting("api_key", "")
    return {"ok": True}


# ── Version ───────────────────────────────────────────────────────────────


@app.get("/api/version")
def api_get_version():
    """Return the current running version."""
    import os
    version_file = Path(__file__).resolve().parent.parent / "VERSION"
    version = "dev"
    if version_file.exists():
        version = version_file.read_text().strip()
    build_id = os.environ.get("ASTRO_BUILD_ID", "")
    return {"version": version, "build_id": build_id}


@app.get("/api/version/latest")
def api_get_latest_version():
    """Check Docker Hub for a newer image of marksnyder/astro."""
    import json, os, re, urllib.request

    version_file = Path(__file__).resolve().parent.parent / "VERSION"
    current = "dev"
    if version_file.exists():
        current = version_file.read_text().strip()
    build_id = os.environ.get("ASTRO_BUILD_ID", "")

    try:
        url = "https://hub.docker.com/v2/repositories/marksnyder/astro/tags/?page_size=50&ordering=last_updated"
        req = urllib.request.Request(url, headers={"Accept": "application/json"})
        with urllib.request.urlopen(req, timeout=5) as resp:
            data = json.loads(resp.read().decode())

        tags = data.get("results", [])
        sha_re = re.compile(r"^[0-9a-f]{7,12}$")
        semver_re = re.compile(r"^\d+\.\d+\.\d+$")

        latest_tag = next((t for t in tags if t["name"] == "latest"), None)
        if not latest_tag:
            return {"current": current, "latest": current, "update_available": False, "build_id": build_id}

        latest_digest = latest_tag.get("digest", "")

        # Strategy 1: compare build SHA against the SHA tag that matches
        # the 'latest' digest (most reliable for Docker-based deploys)
        if build_id:
            latest_sha_tag = next(
                (t["name"] for t in tags if sha_re.match(t["name"]) and t.get("digest") == latest_digest),
                None,
            )
            if latest_sha_tag:
                update_available = not build_id.lower().startswith(latest_sha_tag.lower())
                return {
                    "current": current,
                    "latest": current if not update_available else f"{current} ({latest_sha_tag})",
                    "update_available": update_available,
                    "build_id": build_id,
                    "latest_build": latest_sha_tag,
                }

        # Strategy 2: compare semver VERSION against semver Docker tags
        semver_tags = [t["name"] for t in tags if semver_re.match(t["name"])]
        if semver_tags and current != "dev" and semver_re.match(current):
            semver_tags.sort(key=lambda v: tuple(int(x) for x in v.split(".")), reverse=True)
            latest_ver = semver_tags[0]
            update_available = tuple(int(x) for x in latest_ver.split(".")) > tuple(int(x) for x in current.split("."))
            return {
                "current": current,
                "latest": latest_ver,
                "update_available": update_available,
                "build_id": build_id,
            }

        return {"current": current, "latest": current, "update_available": False, "build_id": build_id}
    except Exception as e:
        print(f"[Astro] WARNING: Version check failed: {e}")
        return {"current": current, "latest": current, "update_available": False, "build_id": build_id}


@app.post("/api/reindex")
def api_reindex():
    """Rebuild the entire vector store from the database and document files.

    Call this after a restore to re-create all embeddings.
    """
    import traceback
    from src.store import clear, add_documents as add_docs, upsert_markdown

    counts = {"markdowns": 0, "document_chunks": 0}

    try:
        # 1. Clear existing vector store
        print("[reindex] Clearing vector store...")
        clear()

        # 2. Re-index markdowns
        print("[reindex] Indexing markdowns...")
        for markdown in list_markdowns():
            upsert_markdown(markdown.id, f"{markdown.title}\n\n{markdown.body}", markdown.title, universe_id=markdown.universe_id)
            counts["markdowns"] += 1

        # 3. Re-index documents from the documents/ folder
        print("[reindex] Indexing documents...")
        all_meta = get_all_document_meta()
        if DOCUMENTS_DIR.is_dir():
            for f in DOCUMENTS_DIR.rglob("*"):
                if not f.is_file():
                    continue
                if f.suffix.lower() not in SUPPORTED_EXTENSIONS:
                    continue
                try:
                    rel = str(f.relative_to(DOCUMENTS_DIR))
                    uid = all_meta.get(rel, {}).get("universe_id", 1)
                    docs = load_document(str(f))
                    if docs:
                        for doc in docs:
                            doc.metadata["source"] = str(f)
                        chunks = chunk_documents(docs)
                        add_docs(chunks, universe_id=uid)
                        counts["document_chunks"] += len(chunks)
                except Exception as e:
                    print(f"[reindex] Error processing {f.name}: {e}")

        print(f"[reindex] Done: {counts}")
        return {"ok": True, "reindexed": counts}

    except Exception as e:
        print(f"[reindex] FATAL: {e}")
        traceback.print_exc()
        raise HTTPException(status_code=500, detail=f"Reindex failed: {str(e)}")




# ── Feeds ─────────────────────────────────────────────────────────────────


class FeedRequest(BaseModel):
    title: str
    category_id: Optional[int] = None


class FeedResponse(BaseModel):
    id: int
    title: str
    category_id: Optional[int]
    universe_id: int
    api_key: str
    pinned: bool = False
    created_at: str
    updated_at: str
    post_count: int = 0
    trend_14d: list[int] = []
    avg_14d: float = 0
    days_since_last: Optional[int] = None


class FeedPostResponse(BaseModel):
    id: int
    feed_id: int
    title: str
    content_type: str
    markdown: Optional[str]
    file_path: Optional[str]
    original_filename: Optional[str]
    created_at: str


class PostCommentRequest(BaseModel):
    author: str = "astro"
    content: str


class PostCommentUpdateRequest(BaseModel):
    content: str


class PostCommentResponse(BaseModel):
    id: int
    post_id: int
    author: str
    content: str
    created_at: str
    updated_at: str


@app.get("/api/feeds", response_model=list[FeedResponse])
def api_list_feeds(q: str = "", category_id: Optional[int] = None, universe_id: Optional[int] = None):
    return [feed_to_dict(f) for f in list_feeds(q, category_id, universe_id=universe_id)]


@app.get("/api/feeds/{feed_id}", response_model=FeedResponse)
def api_get_feed(feed_id: int):
    feed = get_feed(feed_id)
    if not feed:
        raise HTTPException(status_code=404, detail="Feed not found")
    return feed_to_dict(feed)


@app.post("/api/feeds", response_model=FeedResponse, status_code=201)
def api_create_feed(req: FeedRequest, universe_id: int = 1):
    feed = create_feed(req.title, req.category_id, universe_id=universe_id)
    return feed_to_dict(feed)


@app.put("/api/feeds/{feed_id}", response_model=FeedResponse)
def api_update_feed(feed_id: int, req: FeedRequest):
    feed = update_feed(feed_id, req.title, req.category_id)
    if not feed:
        raise HTTPException(status_code=404, detail="Feed not found")
    return feed_to_dict(feed)


@app.post("/api/feeds/{feed_id}/move-universe", response_model=FeedResponse)
def api_move_feed_universe(feed_id: int, req: MoveToUniverseRequest):
    if not get_universe(req.universe_id):
        raise HTTPException(status_code=400, detail="Universe not found")
    _validate_move_category(req)
    feed = move_feed_to_universe(feed_id, req.universe_id, req.category_id)
    if not feed:
        raise HTTPException(status_code=404, detail="Feed not found")
    return feed_to_dict(feed)


@app.delete("/api/feeds/{feed_id}")
def api_delete_feed(feed_id: int):
    if not delete_feed(feed_id):
        raise HTTPException(status_code=404, detail="Feed not found")
    return {"ok": True}


@app.put("/api/feeds/{feed_id}/pin")
def api_pin_feed(feed_id: int, pinned: bool = True):
    if not set_feed_pinned(feed_id, pinned):
        raise HTTPException(status_code=404, detail="Feed not found")
    return {"ok": True}


# ── Feed posts ────────────────────────────────────────────────────────


@app.get("/api/feeds/{feed_id}/posts")
def api_list_feed_posts(feed_id: int, q: str = "", page: int = 1, page_size: int = 100):
    if not get_feed(feed_id):
        raise HTTPException(status_code=404, detail="Feed not found")
    page_size = min(page_size, 100)
    posts, total = list_feed_posts(feed_id, q, page=page, page_size=page_size)
    return {
        "posts": [feed_post_to_dict(p) for p in posts],
        "total": total,
        "page": page,
        "page_size": page_size,
        "has_more": (page * page_size) < total,
    }


@app.get("/api/feed-posts/by-category")
def api_list_feed_posts_by_category(category_id: int = None, q: str = "", page: int = 1, page_size: int = 5):
    page_size = min(page_size, 50)
    posts, total = list_feed_posts_by_category(category_id, q, page=page, page_size=page_size)
    return {
        "posts": posts,
        "total": total,
        "page": page,
        "page_size": page_size,
        "has_more": (page * page_size) < total,
    }


@app.post("/api/feed-posts/mark-read")
def api_mark_posts_read(body: dict):
    ids = body.get("ids", [])
    if not isinstance(ids, list):
        raise HTTPException(status_code=400, detail="ids must be a list")
    updated = mark_feed_posts_read(ids)
    return {"ok": True, "updated": updated}


@app.get("/api/feed-posts/unread-counts")
def api_unread_counts(universe_id: Optional[int] = None):
    counts = get_unread_counts_by_category(universe_id)
    recent = get_recent_counts_by_category(universe_id, days=7)
    fmt = lambda d: {str(k) if k is not None else "null": v for k, v in d.items()}
    return {"counts": fmt(counts), "recent_7d": fmt(recent)}


@app.delete("/api/feed-posts/{post_id}")
def api_delete_feed_post(post_id: int):
    if not delete_feed_post(post_id):
        raise HTTPException(status_code=404, detail="Post not found")
    return {"ok": True}


@app.post("/api/feed-posts/{post_id}/to-markdown")
def api_post_to_markdown(post_id: int):
    """Convert a markdown feed post into a markdown note, preserving links and images as Markdown."""
    from markdownify import markdownify as md

    post = get_feed_post(post_id)
    if not post:
        raise HTTPException(status_code=404, detail="Post not found")
    if post.content_type != "markdown":
        raise HTTPException(status_code=400, detail="Only markdown posts can be converted to markdowns")
    feed = get_feed(post.feed_id)
    uid = feed.universe_id if feed else 1
    cat_id = feed.category_id if feed else None
    markdown_body = md(post.markdown or "", heading_style="ATX", bullets="-").strip()
    markdown = create_markdown(post.title, markdown_body, category_id=cat_id, universe_id=uid)
    try:
        upsert_markdown(markdown.id, f"{markdown.title}\n\n{markdown.body}", markdown.title, universe_id=uid)
    except Exception as e:
        print(f"[Astro] WARNING: Failed to upsert markdown {markdown.id} into vector store: {e}")
    return {"ok": True, "markdown_id": markdown.id}


@app.post("/api/feed-posts/{post_id}/to-document")
def api_post_to_document(post_id: int):
    """Copy a file post into the document archive and ingest it."""
    post = get_feed_post(post_id)
    if not post:
        raise HTTPException(status_code=404, detail="Post not found")
    if post.content_type != "file":
        raise HTTPException(status_code=400, detail="Only file posts can be converted to documents")
    if not post.file_path:
        raise HTTPException(status_code=400, detail="Post has no file")
    src_path = FEED_FILES_DIR / post.file_path
    if not src_path.is_file():
        raise HTTPException(status_code=404, detail="Post file missing from disk")
    feed = get_feed(post.feed_id)
    uid = feed.universe_id if feed else 1
    filename = post.original_filename or post.file_path
    ext = Path(filename).suffix.lower()
    if ext not in SUPPORTED_EXTENSIONS:
        raise HTTPException(status_code=400, detail=f"Unsupported file type: {ext}")
    archive_folder = ext.lstrip(".")
    archive_dir = DOCUMENTS_DIR / archive_folder
    archive_dir.mkdir(exist_ok=True)
    dest = archive_dir / filename
    counter = 1
    while dest.exists():
        stem = Path(filename).stem
        dest = archive_dir / f"{stem}_{counter}{ext}"
        counter += 1
    shutil.copy2(str(src_path), str(dest))
    try:
        documents = load_document(str(dest))
        if documents:
            for doc in documents:
                doc.metadata["source"] = str(dest)
            chunks = chunk_documents(documents)
            add_documents(chunks, universe_id=uid)
    except Exception as e:
        print(f"[Astro] WARNING: Failed to ingest post file as document: {e}")
    rel = dest.relative_to(DOCUMENTS_DIR)
    set_document_universe(str(rel), uid)
    delete_feed_post(post_id)
    return {"ok": True, "path": str(rel)}


# ── Post comments ─────────────────────────────────────────────────────────


@app.get("/api/feed-posts/{post_id}/comments", response_model=list[PostCommentResponse])
def api_list_post_comments(post_id: int):
    return [post_comment_to_dict(c) for c in list_post_comments(post_id)]


@app.post("/api/feed-posts/{post_id}/comments", response_model=PostCommentResponse, status_code=201)
def api_create_post_comment(post_id: int, req: PostCommentRequest):
    if not get_feed_post(post_id):
        raise HTTPException(status_code=404, detail="Post not found")
    return post_comment_to_dict(create_post_comment(post_id, req.author.strip() or "astro", req.content.strip()))


@app.put("/api/post-comments/{comment_id}", response_model=PostCommentResponse)
def api_update_post_comment(comment_id: int, req: PostCommentUpdateRequest):
    c = update_post_comment(comment_id, req.content.strip())
    if not c:
        raise HTTPException(status_code=404, detail="Comment not found")
    return post_comment_to_dict(c)


@app.delete("/api/post-comments/{comment_id}")
def api_delete_post_comment(comment_id: int):
    if not delete_post_comment(comment_id):
        raise HTTPException(status_code=404, detail="Comment not found")
    return {"ok": True}


# ── Feed external ingest endpoint ─────────────────────────────────────────


def _ensure_markdown(text: str) -> str:
    """If *text* looks like HTML, convert it to Markdown; otherwise return as-is."""
    import re
    if re.search(r"<(?:p|div|span|h[1-6]|ul|ol|li|table|br|img|a|strong|em)\b", text, re.I):
        from markdownify import markdownify as md
        return md(text, heading_style="ATX", bullets="-").strip()
    return text


@app.post("/api/feeds/{feed_id}/ingest")
async def api_feed_ingest(
    feed_id: int,
    title: str = fastapi.Form(""),
    markdown: Optional[str] = fastapi.Form(None),
    file: Optional[UploadFile] = None,
    x_feed_key: Optional[str] = fastapi.Header(None),
):
    """External endpoint for pushing posts into a feed.

    Authenticate with the X-Feed-Key header matching the feed's api_key.

    Content in the ``markdown`` field can be Markdown or HTML.  If HTML is
    detected it is automatically converted to Markdown before storage.

    To send markdown (Markdown or HTML):
      POST /api/feeds/{id}/ingest
      Content-Type: multipart/form-data
      X-Feed-Key: fk_...
      title=...&markdown=...

    To send a file:
      POST /api/feeds/{id}/ingest
      Content-Type: multipart/form-data
      X-Feed-Key: fk_...
      title=...
      file=@document.pdf
    """
    feed = get_feed(feed_id)
    if not feed:
        raise HTTPException(status_code=404, detail="Feed not found")

    if not x_feed_key:
        raise HTTPException(status_code=401, detail="Missing X-Feed-Key header")
    if x_feed_key != feed.api_key:
        raise HTTPException(status_code=403, detail="Invalid API key")

    if not title.strip():
        title = "Untitled post"

    if file and file.filename:
        data = await file.read()
        post = create_feed_post_file(feed_id, title.strip(), file.filename, data)
        return {"ok": True, "post_id": post.id, "content_type": "file"}
    elif markdown is not None:
        post = create_feed_post_markdown(feed_id, title.strip(), _ensure_markdown(markdown))
        return {"ok": True, "post_id": post.id, "content_type": "markdown"}
    else:
        raise HTTPException(status_code=400, detail="Provide either 'markdown' or 'file'")


@app.get("/api/feed-files/{filename}")
def api_serve_feed_file(filename: str):
    safe = (FEED_FILES_DIR / filename).resolve()
    if not str(safe).startswith(str(FEED_FILES_DIR.resolve())) or not safe.is_file():
        raise HTTPException(status_code=404, detail="File not found")
    return FileResponse(safe)


class IrcSendRequest(BaseModel):
    message: str


@app.post("/api/irc/send")
def api_irc_send(req: IrcSendRequest):
    from src.irc_client import IRCClient
    client = IRCClient.get()
    client.send_message(req.message)
    return {"ok": True}


@app.get("/api/irc/messages")
def api_irc_messages(after: int = 0):
    from src.irc_client import IRCClient
    client = IRCClient.get()
    return {"messages": client.get_messages(after)}


@app.get("/api/irc/history")
def api_irc_history(channel: str = "#astro", before_id: Optional[int] = None, limit: int = 100):
    """Return persisted IRC history for a channel, paginated by message id."""
    from src.irc_monitor import get_history
    limit = min(limit, 200)
    messages = get_history(channel, before_id=before_id, limit=limit)
    has_more = len(messages) == limit
    return {"messages": messages, "has_more": has_more}


@app.post("/api/irc/unread")
def api_irc_unread(since: dict[str, float]):
    """Return unread message counts per channel since given timestamps."""
    from src.irc_monitor import get_unread_counts
    return get_unread_counts(since)


@app.get("/api/irc/status")
def api_irc_status():
    from src.irc_client import IRCClient
    client = IRCClient.get()
    return client.get_status()


@app.get("/api/irc/users")
def api_irc_users():
    from src.irc_client import IRCClient
    client = IRCClient.get()
    return client.get_channel_users()


@app.get("/api/irc/channels")
def api_irc_channels():
    from src.irc_monitor import IRCMonitor
    return IRCMonitor.get().get_channels()


class IrcChannelRequest(BaseModel):
    name: str


@app.post("/api/irc/channels", status_code=201)
def api_create_irc_channel(req: IrcChannelRequest):
    """Create a channel by having the Astro client join it (IRC creates on join)."""
    from src.irc_client import IRCClient
    name = req.name.strip()
    if not name.startswith("#"):
        name = "#" + name
    client = IRCClient.get()
    if not client.connected or not client._sock:
        raise HTTPException(status_code=503, detail="IRC not connected")
    old_channel = client.channel
    client._raw_send(f"JOIN {name}")
    import time
    time.sleep(0.5)
    client._raw_send(f"PART {name}")
    time.sleep(0.3)
    if old_channel != name:
        client._raw_send(f"JOIN {old_channel}")
    return {"ok": True, "name": name}


@app.post("/api/irc/channels/{name:path}/hide")
def api_hide_irc_channel(name: str):
    """Have both bots leave a channel immediately when it's hidden."""
    from src.irc_client import IRCClient
    from src.irc_monitor import IRCMonitor
    name = name.strip()
    if not name.startswith("#"):
        name = "#" + name
    try:
        client = IRCClient.get()
        if client.connected and client._sock:
            client._raw_send(f"PART {name}")
    except Exception:
        pass
    try:
        IRCMonitor.get().part_channel(name)
    except Exception:
        pass
    return {"ok": True}


@app.delete("/api/irc/channels/{name:path}/history")
def api_purge_irc_channel_history(name: str):
    """Delete all persisted message history for a channel."""
    from src.irc_monitor import purge_history
    name = name.strip()
    if not name.startswith("#"):
        name = "#" + name
    count = purge_history(name)
    return {"ok": True, "deleted": count}


@app.delete("/api/irc/channels/{name:path}")
def api_delete_irc_channel(name: str):
    """Fully delete a channel: have bots leave, remove from ngircd config, purge DB history."""
    import subprocess, re, time
    from src.irc_client import IRCClient
    from src.irc_monitor import IRCMonitor, delete_channel
    name = name.strip()
    if not name.startswith("#"):
        name = "#" + name

    # Have both IRC bots leave the channel first so ngircd can destroy it
    try:
        client = IRCClient.get()
        if client.connected and client._sock:
            client._raw_send(f"PART {name}")
    except Exception:
        pass
    try:
        IRCMonitor.get().part_channel(name)
    except Exception:
        pass
    time.sleep(0.5)

    delete_channel(name)
    conf_path = Path(__file__).resolve().parent.parent / "config" / "ngircd.conf"
    if conf_path.exists():
        text = conf_path.read_text()
        pattern = r'\[Channel\]\s*\n\s*Name\s*=\s*' + re.escape(name) + r'[^\[]*'
        new_text = re.sub(pattern, '', text, flags=re.IGNORECASE)
        if new_text != text:
            conf_path.write_text(new_text.strip() + "\n")
            subprocess.run(["pkill", "-HUP", "ngircd"], capture_output=True)
    return {"ok": True}


@app.put("/api/irc/switch")
def api_irc_switch_channel(req: IrcChannelRequest):
    from src.irc_client import IRCClient
    name = req.name.strip()
    if not name.startswith("#"):
        name = "#" + name
    client = IRCClient.get()
    client.switch_channel(name)
    set_setting("irc_channel", name)
    return {"ok": True, "channel": name}


@app.websocket("/ws/irc")
async def ws_irc(ws: WebSocket):
    import asyncio, json
    from src.irc_client import IRCClient

    await ws.accept()
    client = IRCClient.get()
    queue = client.subscribe()

    # Send current status (history is loaded from DB via /api/irc/history)
    await ws.send_json({"type": "status", **client.get_status()})

    async def _reader():
        """Read send commands from browser."""
        try:
            while True:
                data = await ws.receive_json()
                if data.get("type") == "send" and data.get("message"):
                    client.send_message(data["message"])
        except (WebSocketDisconnect, Exception):
            pass

    async def _writer():
        """Push IRC messages to browser as they arrive."""
        try:
            while True:
                msg = await queue.get()
                await ws.send_json({"type": "msg", **msg})
        except (WebSocketDisconnect, Exception):
            pass

    async def _status():
        """Push status updates periodically."""
        try:
            last = None
            while True:
                await asyncio.sleep(3)
                s = client.get_status()
                if s != last:
                    await ws.send_json({"type": "status", **s})
                    last = s
        except (WebSocketDisconnect, Exception):
            pass

    try:
        await asyncio.gather(_reader(), _writer(), _status())
    except Exception:
        pass
    finally:
        client.unsubscribe(queue)
        try:
            await ws.close()
        except Exception:
            pass


# ── Diagrams ──────────────────────────────────────────────────────────────


class DiagramRequest(BaseModel):
    title: str
    data: str = '{"type":"excalidraw","version":2,"source":"https://excalidraw.com","elements":[],"appState":{"viewBackgroundColor":"#ffffff","gridSize":20},"files":{}}'
    category_id: Optional[int] = None


class DiagramResponse(BaseModel):
    id: int
    title: str
    data: str
    category_id: Optional[int]
    pinned: bool
    created_at: str
    updated_at: str
    universe_id: int = 1


class DiagramSummaryResponse(BaseModel):
    id: int
    title: str
    category_id: Optional[int]
    pinned: bool
    created_at: str
    updated_at: str
    universe_id: int = 1


@app.get("/api/diagrams", response_model=list[DiagramSummaryResponse])
def api_list_diagrams(q: str = "", category_id: Optional[int] = None, universe_id: Optional[int] = None):
    return [diagram_summary_to_dict(d) for d in list_diagram_summaries(q, category_id, universe_id=universe_id)]


@app.get("/api/diagrams/{diagram_id}", response_model=DiagramResponse)
def api_get_diagram(diagram_id: int):
    diagram = get_diagram(diagram_id)
    if not diagram:
        raise HTTPException(status_code=404, detail="Diagram not found")
    return diagram_to_dict(diagram)


@app.post("/api/diagrams", response_model=DiagramResponse, status_code=201)
def api_create_diagram(req: DiagramRequest, universe_id: int = 1):
    diagram = create_diagram(req.title, req.data, req.category_id, universe_id=universe_id)
    return diagram_to_dict(diagram)


@app.put("/api/diagrams/{diagram_id}", response_model=DiagramResponse)
def api_update_diagram(diagram_id: int, req: DiagramRequest):
    diagram = update_diagram(diagram_id, req.title, req.data, req.category_id)
    if not diagram:
        raise HTTPException(status_code=404, detail="Diagram not found")
    return diagram_to_dict(diagram)


@app.post("/api/diagrams/{diagram_id}/move-universe", response_model=DiagramResponse)
def api_move_diagram_universe(diagram_id: int, req: MoveToUniverseRequest):
    if not get_universe(req.universe_id):
        raise HTTPException(status_code=400, detail="Universe not found")
    _validate_move_category(req)
    diagram = move_diagram_to_universe(diagram_id, req.universe_id, req.category_id)
    if not diagram:
        raise HTTPException(status_code=404, detail="Diagram not found")
    return diagram_to_dict(diagram)


@app.delete("/api/diagrams/{diagram_id}")
def api_delete_diagram(diagram_id: int):
    if not delete_diagram(diagram_id):
        raise HTTPException(status_code=404, detail="Diagram not found")
    return {"ok": True}


@app.put("/api/diagrams/{diagram_id}/pin")
def api_toggle_diagram_pin(diagram_id: int, pinned: bool = True):
    if not set_diagram_pinned(diagram_id, pinned):
        raise HTTPException(status_code=404, detail="Diagram not found")
    return {"ok": True}


# ── Tables ────────────────────────────────────────────────────────────────


class TableRequest(BaseModel):
    title: str
    columns: str = "[]"
    category_id: Optional[int] = None


class TableResponse(BaseModel):
    id: int
    title: str
    columns: str
    category_id: Optional[int]
    pinned: bool
    created_at: str
    updated_at: str
    universe_id: int = 1


class TableRowRequest(BaseModel):
    data: str = "{}"
    sort_order: int = 0


class TableRowResponse(BaseModel):
    id: int
    table_id: int
    data: str
    sort_order: int
    created_at: str


@app.get("/api/tables", response_model=list[TableResponse])
def api_list_tables(q: str = "", category_id: Optional[int] = None, universe_id: Optional[int] = None):
    return [table_to_dict(t) for t in list_tables(q, category_id, universe_id=universe_id)]


@app.get("/api/tables/{table_id}", response_model=TableResponse)
def api_get_table(table_id: int):
    t = get_table(table_id)
    if not t:
        raise HTTPException(status_code=404, detail="Table not found")
    return table_to_dict(t)


@app.post("/api/tables", response_model=TableResponse, status_code=201)
def api_create_table(req: TableRequest, universe_id: int = 1):
    t = create_table(req.title, req.columns, req.category_id, universe_id=universe_id)
    return table_to_dict(t)


@app.put("/api/tables/{table_id}", response_model=TableResponse)
def api_update_table(table_id: int, req: TableRequest):
    t = update_table(table_id, req.title, req.columns, req.category_id)
    if not t:
        raise HTTPException(status_code=404, detail="Table not found")
    return table_to_dict(t)


@app.post("/api/tables/{table_id}/move-universe", response_model=TableResponse)
def api_move_table_universe(table_id: int, req: MoveToUniverseRequest):
    if not get_universe(req.universe_id):
        raise HTTPException(status_code=400, detail="Universe not found")
    _validate_move_category(req)
    t = move_table_to_universe(table_id, req.universe_id, req.category_id)
    if not t:
        raise HTTPException(status_code=404, detail="Table not found")
    return table_to_dict(t)


@app.delete("/api/tables/{table_id}")
def api_delete_table(table_id: int):
    if not delete_table(table_id):
        raise HTTPException(status_code=404, detail="Table not found")
    return {"ok": True}


@app.put("/api/tables/{table_id}/pin")
def api_toggle_table_pin(table_id: int, pinned: bool = True):
    if not set_table_pinned(table_id, pinned):
        raise HTTPException(status_code=404, detail="Table not found")
    return {"ok": True}


# ── Table rows ────────────────────────────────────────────────────────────


@app.get("/api/tables/{table_id}/rows")
def api_list_table_rows(table_id: int, search: str = "", page: int = 1, page_size: int = 50):
    if not get_table(table_id):
        raise HTTPException(status_code=404, detail="Table not found")
    page_size = min(page_size, 200)
    rows, total = list_table_rows(table_id, search, page=page, page_size=page_size)
    return {
        "rows": [table_row_to_dict(r) for r in rows],
        "total": total,
        "page": page,
        "page_size": page_size,
        "has_more": (page * page_size) < total,
    }


@app.post("/api/tables/{table_id}/rows", status_code=201)
def api_create_table_row(table_id: int, req: TableRowRequest):
    if not get_table(table_id):
        raise HTTPException(status_code=404, detail="Table not found")
    row = create_table_row(table_id, req.data, req.sort_order)
    return table_row_to_dict(row)


@app.put("/api/table-rows/{row_id}")
def api_update_table_row(row_id: int, req: TableRowRequest):
    row = update_table_row(row_id, req.data, req.sort_order)
    if not row:
        raise HTTPException(status_code=404, detail="Row not found")
    return table_row_to_dict(row)


@app.delete("/api/table-rows/{row_id}")
def api_delete_table_row(row_id: int):
    if not delete_table_row(row_id):
        raise HTTPException(status_code=404, detail="Row not found")
    return {"ok": True}


# ── Table CSV import/export ───────────────────────────────────────────────


@app.get("/api/tables/{table_id}/export-csv")
def api_export_table_csv(table_id: int):
    import csv, io, json
    t = get_table(table_id)
    if not t:
        raise HTTPException(status_code=404, detail="Table not found")
    columns = json.loads(t.columns)
    col_names = [c["name"] for c in columns]
    all_rows, _ = list_table_rows(table_id, page=1, page_size=100000)
    output = io.StringIO()
    writer = csv.writer(output)
    writer.writerow(col_names)
    for row in all_rows:
        row_data = json.loads(row.data)
        writer.writerow([row_data.get(cn, "") for cn in col_names])
    from fastapi.responses import Response
    return Response(
        content=output.getvalue(),
        media_type="text/csv",
        headers={"Content-Disposition": f'attachment; filename="{t.title or "table"}.csv"'},
    )


@app.post("/api/tables/{table_id}/import-csv")
async def api_import_table_csv(table_id: int, file: UploadFile):
    import csv, io, json
    t = get_table(table_id)
    if not t:
        raise HTTPException(status_code=404, detail="Table not found")
    columns = json.loads(t.columns)
    col_map = {c["name"]: c["type"] for c in columns}
    content = (await file.read()).decode("utf-8-sig")
    reader = csv.DictReader(io.StringIO(content))
    count = 0
    for csv_row in reader:
        data = {}
        for key, val in csv_row.items():
            if key not in col_map:
                continue
            ctype = col_map[key]
            if ctype == "number":
                try:
                    data[key] = float(val) if val else 0
                except ValueError:
                    data[key] = 0
            elif ctype == "boolean":
                data[key] = val.lower() in ("true", "1", "yes") if val else False
            else:
                data[key] = val or ""
        create_table_row(table_id, json.dumps(data), count)
        count += 1
    return {"ok": True, "imported": count}


@app.post("/api/tables/import-csv-new")
async def api_import_csv_new_table(file: UploadFile, universe_id: int = 1):
    import csv, io, json
    if not file.filename:
        raise HTTPException(status_code=400, detail="No filename")
    content = (await file.read()).decode("utf-8-sig")
    reader = csv.DictReader(io.StringIO(content))
    fieldnames = reader.fieldnames or []
    if not fieldnames:
        raise HTTPException(status_code=400, detail="CSV has no headers")
    # Infer column types from first row
    first_rows = []
    for i, row in enumerate(reader):
        first_rows.append(row)
        if i >= 20:
            break
    columns = []
    for fn in fieldnames:
        ctype = "string"
        sample_vals = [r.get(fn, "") for r in first_rows if r.get(fn, "")]
        if sample_vals:
            if all(v.lower() in ("true", "false", "1", "0", "yes", "no") for v in sample_vals):
                ctype = "boolean"
            else:
                try:
                    for v in sample_vals:
                        float(v)
                    ctype = "number"
                except ValueError:
                    pass
        columns.append({"name": fn, "type": ctype})
    title = file.filename.rsplit(".", 1)[0] if "." in file.filename else file.filename
    t = create_table(title, json.dumps(columns), universe_id=universe_id)
    # Re-read for all rows
    all_reader = csv.DictReader(io.StringIO(content))
    col_map = {c["name"]: c["type"] for c in columns}
    count = 0
    for csv_row in all_reader:
        data = {}
        for key, val in csv_row.items():
            if key not in col_map:
                continue
            ctype = col_map[key]
            if ctype == "number":
                try:
                    data[key] = float(val) if val else 0
                except ValueError:
                    data[key] = 0
            elif ctype == "boolean":
                data[key] = val.lower() in ("true", "1", "yes") if val else False
            else:
                data[key] = val or ""
        create_table_row(t.id, json.dumps(data), count)
        count += 1
    return {"ok": True, "table_id": t.id, "title": t.title, "columns": len(columns), "rows": count}


# ── Search ────────────────────────────────────────────────────────────────


@app.get("/api/search")
def api_search(q: str, k: int = 4, universe_id: int = 1):
    """Semantic search over the vector store. Returns the top-k most relevant
    chunks for the given query, useful for RAG pipelines and agent tooling."""
    if not q.strip():
        raise HTTPException(status_code=400, detail="Query parameter 'q' is required")
    k = max(1, min(k, 20))
    retriever = get_retriever(k=k, universe_id=universe_id)
    docs = retriever.invoke(q)
    return {
        "query": q,
        "universe_id": universe_id,
        "results": [
            {
                "content": d.page_content,
                "source": d.metadata.get("source", ""),
                "metadata": {k: v for k, v in d.metadata.items() if k != "source"},
            }
            for d in docs
        ],
    }


# ── Stats ────────────────────────────────────────────────────────────────


@app.get("/api/stats", response_model=StatsResponse)
def api_stats():
    from src.migrate import get_current_version
    from src.markdowns import _get_conn
    conn = _get_conn()
    version = get_current_version(conn)
    conn.close()
    return StatsResponse(chunks=doc_count(), schema_version=version)


# Serve built React app if it exists
_static = Path(__file__).resolve().parent.parent / "web" / "dist"
if _static.is_dir():
    # SPA catch-all: serve index.html for client-side routes like /mobile
    _index_html = _static / "index.html"

    @app.get("/sw.js")
    def service_worker():
        return FileResponse(
            str(_static / "sw.js"),
            media_type="application/javascript",
            headers={"Service-Worker-Allowed": "/", "Cache-Control": "no-cache"},
        )

    @app.get("/manifest.json")
    def pwa_manifest():
        return FileResponse(
            str(_static / "manifest.json"),
            media_type="application/manifest+json",
        )

    @app.get("/mobile")
    @app.get("/mobile/{rest:path}")
    def spa_mobile(rest: str = ""):
        return FileResponse(str(_index_html))

    app.mount("/", StaticFiles(directory=str(_static), html=True), name="static")
